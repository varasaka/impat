"""
IMPPAT Phytochemical Explorer (GUI)
=====================================

A simple desktop app: type a plant name, hit Enter (or click Fetch),
and it scrapes IMPPAT for every phytochemical of that plant — name,
SMILES, molecular formula, and molecular weight — then saves the
result as an Excel (.xlsx) file next to the script.

Requirements:
    pip install requests beautifulsoup4 openpyxl --break-system-packages

Run:
    python imppat_gui.py

(Run this from a terminal, or just double-click it — this version
uses a proper window so it won't vanish like a console script does.)
"""

import re
import sys
import time
import threading
import urllib.parse
from pathlib import Path
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# ------------------------------------------------------------------
# Scraping logic (same as the CLI tool)
# ------------------------------------------------------------------

BASE = "https://cb.imsc.res.in/imppat"
PLANT_URL_TMPL = f"{BASE}/phytochemical/{{plant}}"
DETAIL_URL_TMPL = f"{BASE}/phytochemical-detailedpage/{{cid}}"
PHYSCHEM_URL_TMPL = f"{BASE}/physicochemicalproperties/{{cid}}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

DELAY = 0.4
INCHI_FORMULA_RE = re.compile(r"InChI=1S?/([A-Za-z0-9.]+)/")


def get_soup(url: str) -> BeautifulSoup:
    resp = SESSION.get(url, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def fetch_compound_list(plant_name: str, log):
    plant_url = PLANT_URL_TMPL.format(plant=urllib.parse.quote(plant_name))
    soup = get_soup(plant_url)

    table = soup.find("table", id="table_id") or soup.find("table")
    if table is None:
        log(f"No phytochemical table found for '{plant_name}'. "
            f"Check the spelling matches IMPPAT exactly.")
        return []

    compounds = {}
    for row in table.find("tbody").find_all("tr"):
        cells = row.find_all("td")
        if len(cells) < 4:
            continue
        id_cell, name_cell = cells[2], cells[3]
        id_link = id_cell.find("a")
        name_link = name_cell.find("a")
        if not id_link or not name_link:
            continue
        cid = id_link.get_text(strip=True)
        name = name_link.get_text(strip=True)
        if cid and name and cid not in compounds:
            compounds[cid] = name

    info_div = soup.find(id="table_id_info")
    if info_div:
        m = re.search(r"of\s+([\d,]+)\s+entries", info_div.get_text())
        if m:
            total_expected = int(m.group(1).replace(",", ""))
            if len(compounds) < total_expected:
                log(f"Warning: page reports {total_expected} total entries but "
                    f"only {len(compounds)} were found — results may be incomplete.")

    return list(compounds.items())


def extract_smiles_and_inchi(soup: BeautifulSoup):
    smiles, inchi = "", ""
    for strong in soup.find_all("strong"):
        label = strong.get_text(strip=True)
        if label == "SMILES:":
            text_tag = strong.find_next("text")
            if text_tag:
                smiles = text_tag.get_text(strip=True)
        elif label == "InChI:":
            text_tag = strong.find_next("text")
            if text_tag:
                inchi = text_tag.get_text(strip=True)
    return smiles, inchi


def formula_from_inchi(inchi: str) -> str:
    if not inchi:
        return ""
    m = INCHI_FORMULA_RE.search(inchi)
    return m.group(1) if m else ""


def extract_mw(soup: BeautifulSoup) -> str:
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if not cells:
            continue
        label = cells[0].get_text(strip=True).lower()
        if "molecular weight" in label:
            value_cell = cells[-1].get_text(strip=True)
            m = re.search(r"[\d]+\.?\d*", value_cell)
            if m:
                return m.group(0)
    return ""


def scrape_compound(cid: str, name: str) -> dict:
    row = {"Compound_ID": cid, "Name": name, "SMILES": "", "Formula": "", "MW_g_per_mol": ""}
    try:
        detail_soup = get_soup(DETAIL_URL_TMPL.format(cid=cid))
        smiles, inchi = extract_smiles_and_inchi(detail_soup)
        row["SMILES"] = smiles
        row["Formula"] = formula_from_inchi(inchi)
    except Exception:
        pass
    time.sleep(DELAY)
    try:
        phys_soup = get_soup(PHYSCHEM_URL_TMPL.format(cid=cid))
        row["MW_g_per_mol"] = extract_mw(phys_soup)
    except Exception:
        pass
    time.sleep(DELAY)
    return row


# ------------------------------------------------------------------
# Excel writer
# ------------------------------------------------------------------

def write_excel(plant_name: str, rows: list, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Phytochemicals"

    headers = ["Compound ID", "Name", "SMILES", "Formula", "MW (g/mol)"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([
            row.get("Compound_ID", ""),
            row.get("Name", ""),
            row.get("SMILES", ""),
            row.get("Formula", ""),
            row.get("MW_g_per_mol", ""),
        ])

    widths = [16, 34, 60, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)


# ------------------------------------------------------------------
# GUI
# ------------------------------------------------------------------

class ImppatApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("IMPPAT Phytochemical Explorer")
        self.geometry("640x480")
        self.minsize(560, 420)

        self.outdir = Path.cwd()

        # --- input row ---
        top = ttk.Frame(self, padding=12)
        top.pack(fill="x")

        ttk.Label(top, text="Plant name:").pack(side="left")
        self.entry = ttk.Entry(top)
        self.entry.pack(side="left", fill="x", expand=True, padx=8)
        self.entry.bind("<Return>", lambda e: self.on_fetch())
        self.entry.focus_set()

        self.fetch_btn = ttk.Button(top, text="Fetch", command=self.on_fetch)
        self.fetch_btn.pack(side="left")

        # --- output folder row ---
        folder_row = ttk.Frame(self, padding=(12, 0))
        folder_row.pack(fill="x")
        self.folder_label = ttk.Label(folder_row, text=f"Save to: {self.outdir}")
        self.folder_label.pack(side="left")
        ttk.Button(folder_row, text="Change...", command=self.choose_folder).pack(side="right")

        # --- log box ---
        log_frame = ttk.Frame(self, padding=12)
        log_frame.pack(fill="both", expand=True)

        self.log_box = tk.Text(log_frame, wrap="word", state="disabled", height=18)
        self.log_box.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(log_frame, command=self.log_box.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_box["yscrollcommand"] = scrollbar.set

        # --- status bar ---
        self.status = ttk.Label(self, text="Ready.", padding=(12, 4), anchor="w")
        self.status.pack(fill="x")

    def log(self, msg: str):
        self.log_box.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def choose_folder(self):
        chosen = filedialog.askdirectory(initialdir=str(self.outdir))
        if chosen:
            self.outdir = Path(chosen)
            self.folder_label.config(text=f"Save to: {self.outdir}")

    def on_fetch(self):
        plant_name = self.entry.get().strip()
        if not plant_name:
            messagebox.showwarning("Missing input", "Please type a plant name first.")
            return

        self.fetch_btn.config(state="disabled")
        self.status.config(text=f"Fetching data for '{plant_name}'...")
        self.log(f"Starting fetch for '{plant_name}'")

        # Run the scrape in a background thread so the GUI doesn't freeze.
        thread = threading.Thread(target=self.run_scrape, args=(plant_name,), daemon=True)
        thread.start()

    def run_scrape(self, plant_name: str):
        try:
            compounds = fetch_compound_list(plant_name, self.log)
            self.after(0, self.log, f"Found {len(compounds)} unique compounds.")

            rows = []
            for i, (cid, name) in enumerate(compounds, 1):
                rows.append(scrape_compound(cid, name))
                self.after(0, self.log, f"[{i}/{len(compounds)}] {cid}  {name}")
                self.after(0, self.status.config, {"text": f"Fetching {i}/{len(compounds)}: {name}"})

            safe_name = re.sub(r"\s+", "_", plant_name.strip())
            out_path = self.outdir / f"{safe_name}_phytochemicals.xlsx"
            write_excel(plant_name, rows, out_path)

            self.after(0, self.log, f"Saved: {out_path}")
            self.after(0, self.status.config, {"text": f"Done — {len(rows)} compounds saved to {out_path.name}"})
            self.after(0, lambda: messagebox.showinfo(
                "Done", f"Saved {len(rows)} compounds to:\n{out_path}"
            ))
        except requests.exceptions.RequestException as e:
            self.after(0, self.log, f"Network error: {e}")
            self.after(0, lambda: messagebox.showerror("Network error", str(e)))
            self.after(0, self.status.config, {"text": "Failed — see log."})
        except Exception as e:
            self.after(0, self.log, f"Error: {e}")
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
            self.after(0, self.status.config, {"text": "Failed — see log."})
        finally:
            self.after(0, self.fetch_btn.config, {"state": "normal"})


if __name__ == "__main__":
    app = ImppatApp()
    app.mainloop()
